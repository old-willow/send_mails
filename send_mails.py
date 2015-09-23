#!/usr/bin/env python
# -*- conding:utf-8 -*-


from openpyxl import load_workbook
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email import Encoders
import os

import getpass


def send_email(to, subject, text, attach, user, passw):
    gmail_user = user
    gmail_password = passw

    msg = MIMEMultipart()

    msg['subject'] = subject
    msg['From'] = gmail_user
    msg['To'] = to

    msg.attach(MIMEText(text))

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(open(attach, 'rb').read())
    Encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachement; filename="%s"' % os.path.basename(attach))
    msg.attach(part)

    # server stuff
    mailServer = smtplib.SMTP('smtp.gmail.com', 587)
    mailServer.ehlo()
    mailServer.starttls()
    mailServer.ehlo()
    mailServer.login(gmail_user, gmail_password)
    mailServer.sendmail(gmail_user, to, msg.as_string())
    mailServer.close()


def fetch_emails_from_excel(file):
    wb = load_workbook(file)
    sheet = wb.get_active_sheet()

    contacts = {}
    counter = 0

    for row in sheet.iter_rows(range_string='A2:C5'):
        print unicode(row[0].value) + ' ' + unicode(row[1].value) + ' ' + unicode(row[2].value)
        recipiant = 'recipient_%03d' % counter
        #contacts[recipiant] = [row[0].value, row[1].value, row[2].value]
        contacts[recipiant] = row[2].value
        counter += 1

    for cont in contacts.items():
        print cont

    return contacts


def main():
    file_name = 'addressbook.xlsx'
    attached_file = 'attachement-test-file.doc'

    contacts = fetch_emails_from_excel(file_name)
    #for c in contacts.items():
    #    print c[1]

    subject = '[EMAIL - TEST] Testing sending email from python program'
    text_body = 'This is a test text body!'

    #passw = raw_input('Password: ')
    user = raw_input('Gmail account email: ')
    passw = getpass.getpass(prompt='Type the password for your gmail account: ')

    for to in contacts.items():
        send_email(to[1], subject, text_body, attached_file, user, passw)


if __name__ == '__main__':
    main()
